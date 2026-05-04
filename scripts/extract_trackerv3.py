#!/usr/bin/env python3
"""
Extract trackerv3.xlsx into a normalized JSON the TS importer can consume.

Usage:
    python3 scripts/extract_trackerv3.py \
        --xlsx "/Users/nomads/tracker/new design/trackerv3.xlsx" \
        --out scripts/trackerv3_import.json
"""
import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl not installed. Run: pip3 install openpyxl\n")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

STANCE_MAP = {
    "support": "support",
    "oppose": "oppose",
    "mixed": "mixed",
    "unclear": "unclear",
    "no mention": "no_mention",
    "no_mention": "no_mention",
    "": "no_mention",
    None: "no_mention",
}

CONFIDENCE_MAP = {
    "high": "high",
    "medium": "medium",
    "low": "low",
    "n/a": "medium",
    "": "medium",
    None: "medium",
}

PARTY_MAP = {
    "democratic": "D",
    "democrat": "D",
    "d": "D",
    "republican": "R",
    "r": "R",
    "independent": "I",
    "i": "I",
    "libertarian": "L",
    "l": "L",
    "green": "G",
    "g": "G",
}

SOURCE_TYPE_MAP = {
    "statement": "statement",
    "legislation": "legislation",
    "vote": "legislation",
    "committee hearing": "statement",
    "news": "news",
    "social media": "social_media",
    "survey": "survey",
}


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def norm_stance(s):
    key = (s or "").strip().lower()
    return STANCE_MAP.get(key, "no_mention")


def norm_confidence(c):
    key = (c or "").strip().lower()
    return CONFIDENCE_MAP.get(key, "medium")


def norm_party(p):
    key = (p or "").strip().lower()
    if not key:
        return "I"  # fallback
    return PARTY_MAP.get(key, "I")


def norm_source_type(t):
    key = (t or "").strip().lower()
    return SOURCE_TYPE_MAP.get(key, "statement")


def split_name(full_name):
    """Split a display name into (first, last). Best-effort."""
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def slugify(s):
    s = (s or "").strip().lower()
    s = re.sub(r"['\"]", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def chamber_and_district(seat, district):
    """Map ('Senate', None/'') -> ('senate', None). ('House', 'CD 3') -> ('house', '03')."""
    seat_norm = (seat or "").strip().lower()
    if seat_norm == "senate":
        return "senate", None
    # House
    if district is None:
        d = ""
    elif isinstance(district, (int, float)):
        d = str(int(district))
    else:
        d = str(district).strip()
    m = re.search(r"(\d+)", d)
    if m:
        return "house", m.group(1).zfill(2)
    return "house", None


def race_slug(state, chamber, district, year=2026):
    state = (state or "").lower()
    if chamber == "senate":
        return f"{state}-sen-{year}"
    return f"{state}-house-{district or '00'}-{year}"


def office_label(chamber, district):
    if chamber == "senate":
        return "U.S. Senate"
    return f"U.S. House, District {int(district)}" if district else "U.S. House"


# ---------------------------------------------------------------------------
# Extract each sheet
# ---------------------------------------------------------------------------


def extract_issues(wb):
    ws = wb["Topics"]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]:
            continue
        slug = clean(row[0])
        display = clean(row[1])
        # shortName: row[2] — not yet used by the schema
        description = clean(row[3]) or ""
        rows.append(
            {
                "slug": slug,
                "display_name": display,
                "description": description,
                "sort_order": i,
            }
        )
    return rows


def extract_candidates(wb):
    ws = wb["Candidates"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        slug_base = clean(row[0])
        name = clean(row[1])
        state = (clean(row[2]) or "").upper()
        party = norm_party(row[3])
        seat = clean(row[4])
        district_raw = clean(row[5])
        is_incumbent = (clean(row[6]) or "").upper() == "Y"
        amount_raised = row[7]
        if isinstance(amount_raised, (int, float)):
            total_raised = int(amount_raised)
        else:
            total_raised = None

        chamber, district = chamber_and_district(seat, district_raw)
        first, last = split_name(name)
        cand_slug = f"{slug_base}-{state.lower()}"

        rows.append(
            {
                "sheet_id": slug_base,
                "slug": cand_slug,
                "name": name,
                "first_name": first,
                "last_name": last,
                "party": party,
                "state": state,
                "district": district,
                "office_sought": office_label(chamber, district),
                "is_incumbent": is_incumbent,
                "total_raised": total_raised,
                "chamber": chamber,
                # race slug this candidate belongs to
                "race_slug": race_slug(state, chamber, district),
            }
        )
    return rows


def extract_positions(wb):
    """Pull first source URL per position out of the Sources sheet."""
    # Build position_id -> first source
    srcs = wb["Sources"]
    source_index = {}  # positionId -> {"url", "type", "date"}
    for row in srcs.iter_rows(min_row=2, values_only=True):
        pos_id = clean(row[0])
        if not pos_id:
            continue
        url = clean(row[3])
        s_type = norm_source_type(row[1])
        s_date_raw = row[4]
        if s_date_raw is None:
            s_date = None
        elif isinstance(s_date_raw, str):
            s_date = clean(s_date_raw)
        elif hasattr(s_date_raw, "isoformat"):
            s_date = s_date_raw.isoformat()[:10]
        else:
            s_date = None

        # Keep first source only (the sheet is in a stable order per positionId)
        if pos_id not in source_index:
            source_index[pos_id] = {"url": url, "type": s_type, "date": s_date}

    ws = wb["Positions v2"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pos_id = clean(row[0])
        cand_id = clean(row[1])
        topic_id = clean(row[2])
        if not pos_id or not cand_id or not topic_id:
            continue
        stance = norm_stance(row[3])
        confidence = norm_confidence(row[4])
        summary = clean(row[5])
        last_updated = row[6]
        if last_updated is None:
            last_updated = None
        elif isinstance(last_updated, str):
            last_updated = clean(last_updated)
        elif hasattr(last_updated, "isoformat"):
            last_updated = last_updated.isoformat()[:10]
        else:
            last_updated = None

        src = source_index.get(pos_id)
        source_url = src["url"] if src else None
        date_recorded = src["date"] if src else None

        rows.append(
            {
                "position_sheet_id": pos_id,
                "candidate_sheet_id": cand_id,
                "issue_slug": topic_id,
                "stance": stance,
                "confidence": confidence,
                "summary": summary,
                "source_url": source_url,
                "date_recorded": date_recorded,
                "last_updated": last_updated,
            }
        )
    return rows


def derive_races(candidates):
    seen = {}
    for c in candidates:
        slug = c["race_slug"]
        if slug in seen:
            continue
        seen[slug] = {
            "slug": slug,
            "state": c["state"],
            "chamber": c["chamber"],
            "district": c["district"],
            "election_year": 2026,
            "race_type": "regular",
        }
    return list(seen.values())


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    issues = extract_issues(wb)
    candidates = extract_candidates(wb)
    positions = extract_positions(wb)
    races = derive_races(candidates)

    # Sanity counts
    cand_ids = {c["sheet_id"]: c["slug"] for c in candidates}
    issue_slugs = {i["slug"] for i in issues}

    orphan_positions = [
        p
        for p in positions
        if p["candidate_sheet_id"] not in cand_ids
        or p["issue_slug"] not in issue_slugs
    ]
    if orphan_positions:
        print(
            f"WARNING: {len(orphan_positions)} positions reference unknown candidate/topic ids; dropping",
            file=sys.stderr,
        )
        positions = [p for p in positions if p not in orphan_positions]

    # Attach candidate_slug to each position for the TS importer
    for p in positions:
        p["candidate_slug"] = cand_ids[p["candidate_sheet_id"]]

    out = {
        "issues": issues,
        "candidates": candidates,
        "races": races,
        "positions": positions,
        "_stats": {
            "issues": len(issues),
            "candidates": len(candidates),
            "races": len(races),
            "positions": len(positions),
            "positions_with_source": sum(1 for p in positions if p.get("source_url")),
            "positions_coded": sum(1 for p in positions if p["stance"] != "no_mention"),
        },
    }

    Path(args.out).write_text(json.dumps(out, indent=2, default=str))
    print(json.dumps(out["_stats"], indent=2))


if __name__ == "__main__":
    main()
