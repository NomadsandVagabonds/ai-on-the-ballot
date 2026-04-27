#!/usr/bin/env python3
"""
Validate research subagent output and stage it into trackerv3.xlsx as
DRAFT rows. The xlsx is the human review surface; only rows a human
flips to reviewStatus="approved" ever reach data/tracker/positions.json
(via build_tracker_json.py). This script never writes positions.json.

For each scripts/.research-cache/{candidate_slug}.json:

  1. Schema-check (stance / confidence enums, required fields per stance,
     exactly 10 issue entries per file).
  2. URL-check every source (concurrent HEAD/GET; HTTP 2xx required).
  3. Quote-check: each `excerpt` and `full_quote` must appear as a
     normalized substring of the live page body. Failed quote check =
     drop the excerpt text but keep the URL. Failed URL = drop source.
     A position that loses its last source reverts to no_mention.
  4. Stage into "new design/trackerv3.xlsx":
       - Positions v2 row updated (stance/confidence/summary/lastUpdated)
         with reviewStatus = "draft"
       - Sources rows for that position id are replaced with the freshly
         verified set (delete-then-append, idempotent re-runs).
  5. Write a per-run report to scripts/.research-cache/_report.md.

Usage:
  python3 scripts/merge_research.py            # validate + stage to xlsx
  python3 scripts/merge_research.py --dry-run  # validate only, no writes
  python3 scripts/merge_research.py jane-doe   # one slug only
"""
from __future__ import annotations

import argparse
import json
import re
import ssl
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

try:
    import certifi  # type: ignore[import-not-found]

    _SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    # macOS system Python ships without a CA bundle — without certifi
    # urlopen will fail certificate verification on most HTTPS hosts.
    _SSL_CTX = ssl.create_default_context()

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "tracker"
CACHE = ROOT / "scripts" / ".research-cache"
ISSUES_PATH = DATA / "issues.json"
CANDIDATES_PATH = DATA / "candidates.json"
XLSX_PATH = ROOT.parent / "new design" / "trackerv3.xlsx"
REPORT_PATH = CACHE / "_report.md"

STANCE_TO_XLSX = {
    "support": "Support",
    "oppose": "Oppose",
    "mixed": "Mixed",
    "unclear": "Unclear",
    "no_mention": "No mention",
}
CONFIDENCE_TO_XLSX = {"high": "High", "medium": "Medium", "low": "Low"}
SOURCE_TYPE_TO_XLSX = {
    "news": "News",
    "statement": "Statement",
    "legislation": "Legislation",
    "survey": "Survey",
    "social_media": "Social Media",
}

VALID_STANCES = {"support", "oppose", "mixed", "unclear", "no_mention"}
VALID_CONFIDENCE = {"high", "medium", "low"}
VALID_SOURCE_TYPES = {"statement", "legislation", "survey", "social_media", "news"}
TODAY = date.today().isoformat()
USER_AGENT = "Mozilla/5.0 (compatible; AIonBallot-research-validator/1.0)"
HTTP_TIMEOUT_SECS = 20


# ----------------------------------------------------------------------
# HTML → text (used for quote substring checks)


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._chunks: list[str] = []
        self._skip = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style", "noscript"}:
            self._skip += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript"} and self._skip > 0:
            self._skip -= 1

    def handle_data(self, data: str) -> None:
        if self._skip == 0:
            self._chunks.append(data)

    def text(self) -> str:
        return " ".join(self._chunks)


_SMART_QUOTES = {
    "‘": "'",
    "’": "'",
    "“": '"',
    "”": '"',
    "′": "'",
    "″": '"',
}
_DASHES = {
    "–": "-",
    "—": "-",
    "−": "-",
    "‐": "-",
    "‑": "-",
}
_OTHER = {
    " ": " ",   # nbsp
    "…": "...",  # ellipsis
}


def _normalize(s: str) -> str:
    """Collapse whitespace, normalize smart punctuation, lowercase.

    The page extractor and the LLM-supplied excerpts use different quote
    styles, hyphens, and whitespace; normalize aggressively before doing
    the substring check so we don't reject otherwise-valid citations.
    """
    for src, dst in _SMART_QUOTES.items():
        s = s.replace(src, dst)
    for src, dst in _DASHES.items():
        s = s.replace(src, dst)
    for src, dst in _OTHER.items():
        s = s.replace(src, dst)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def fetch_page(url: str) -> tuple[bool, str | None, str]:
    """Return (ok, body_text, error_message). Tries GET; HEAD-only servers
    we treat as ok with no body (so quote-check has to rely on metadata)."""
    try:
        req = Request(url, headers={"User-Agent": USER_AGENT})
        with urlopen(req, timeout=HTTP_TIMEOUT_SECS, context=_SSL_CTX) as resp:
            status = getattr(resp, "status", 200)
            if status >= 400:
                return False, None, f"HTTP {status}"
            ctype = resp.headers.get_content_type() or ""
            raw = resp.read(2_000_000)  # cap at 2 MB
            if "html" in ctype or "xml" in ctype:
                parser = _TextExtractor()
                try:
                    parser.feed(raw.decode("utf-8", errors="replace"))
                except Exception:
                    return True, None, ""
                return True, parser.text(), ""
            elif "text" in ctype:
                return True, raw.decode("utf-8", errors="replace"), ""
            else:
                # Non-text content (e.g., PDF): URL is reachable but we
                # can't substring-check; accept the source as long as the
                # excerpt is non-empty.
                return True, None, ""
    except HTTPError as e:
        return False, None, f"HTTP {e.code}"
    except URLError as e:
        return False, None, f"URL error: {e.reason}"
    except Exception as e:  # noqa: BLE001
        return False, None, f"{type(e).__name__}: {e}"


# ----------------------------------------------------------------------
# Validation


@dataclass
class ValidationResult:
    candidate_slug: str
    candidate_id: str | None = None
    coded_count: int = 0
    no_mention_count: int = 0
    rejected_sources: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    positions_to_apply: list[dict[str, Any]] = field(default_factory=list)


def validate_position_entry(
    entry: dict[str, Any],
    issue_slug: str,
    cache: dict[str, tuple[bool, str | None, str]],
) -> tuple[dict[str, Any] | None, list[str]]:
    """Return (cleaned_entry, errors). cleaned_entry is None on hard error."""
    errs: list[str] = []
    stance = entry.get("stance")
    if stance not in VALID_STANCES:
        errs.append(f"[{issue_slug}] invalid stance: {stance!r}")
        return None, errs

    if stance == "no_mention":
        return (
            {
                "stance": "no_mention",
                "confidence": "low",
                "summary": None,
                "full_quote": None,
                "sources": [],
            },
            errs,
        )

    confidence = entry.get("confidence")
    if confidence not in VALID_CONFIDENCE:
        errs.append(f"[{issue_slug}] invalid confidence: {confidence!r}")
        return None, errs

    summary = entry.get("summary")
    if not isinstance(summary, str) or not summary.strip():
        errs.append(f"[{issue_slug}] non-no_mention stance requires a summary")
        return None, errs

    raw_sources = entry.get("sources") or []
    if not isinstance(raw_sources, list) or not raw_sources:
        errs.append(f"[{issue_slug}] non-no_mention stance requires ≥1 source")
        return None, errs

    cleaned_sources: list[dict[str, Any]] = []
    for i, src in enumerate(raw_sources):
        url = (src.get("url") or "").strip()
        if not url or not urlparse(url).scheme.startswith("http"):
            errs.append(f"[{issue_slug}] source #{i}: missing/invalid URL")
            continue
        stype = src.get("type")
        if stype not in VALID_SOURCE_TYPES:
            errs.append(f"[{issue_slug}] source #{i}: invalid type {stype!r}")
            continue

        ok, body, msg = cache.get(url, (False, None, "not fetched"))
        if not ok:
            errs.append(f"[{issue_slug}] source #{i} unreachable ({msg}): {url}")
            continue

        excerpt = (src.get("excerpt") or "").strip() or None
        if excerpt and body is not None:
            if _normalize(excerpt) not in _normalize(body):
                errs.append(
                    f"[{issue_slug}] source #{i}: excerpt not found verbatim on page: {url}"
                )
                # Drop the excerpt but keep the source if URL is live.
                excerpt = None

        cleaned_sources.append(
            {
                "type": stype,
                "title": (src.get("title") or "").strip() or None,
                "url": url,
                "date": (src.get("date") or "").strip() or None,
                "excerpt": excerpt,
            }
        )

    if not cleaned_sources:
        errs.append(f"[{issue_slug}] all sources rejected; demoting to no_mention")
        return (
            {
                "stance": "no_mention",
                "confidence": "low",
                "summary": None,
                "full_quote": None,
                "sources": [],
            },
            errs,
        )

    full_quote = entry.get("full_quote")
    if isinstance(full_quote, str) and not full_quote.strip():
        full_quote = None
    if isinstance(full_quote, str):
        # full_quote should appear in at least one fetched source body.
        norm = _normalize(full_quote)
        in_a_body = any(
            cache[s["url"]][1] is not None
            and norm in _normalize(cache[s["url"]][1] or "")
            for s in cleaned_sources
            if s["url"] in cache
        )
        if not in_a_body:
            errs.append(f"[{issue_slug}] full_quote not verbatim on any source page; clearing it")
            full_quote = None

    return (
        {
            "stance": stance,
            "confidence": confidence,
            "summary": summary.strip(),
            "full_quote": full_quote,
            "sources": cleaned_sources,
        },
        errs,
    )


def collect_urls(payload: dict[str, Any]) -> list[str]:
    urls: list[str] = []
    for p in payload.get("positions") or []:
        for s in p.get("sources") or []:
            u = (s or {}).get("url")
            if u:
                urls.append(u.strip())
    # dedupe, preserve order
    seen: set[str] = set()
    out: list[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def fetch_all(urls: list[str], workers: int = 8) -> dict[str, tuple[bool, str | None, str]]:
    out: dict[str, tuple[bool, str | None, str]] = {}
    if not urls:
        return out
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(fetch_page, u): u for u in urls}
        for fut in as_completed(futures):
            url = futures[fut]
            try:
                out[url] = fut.result()
            except Exception as e:  # noqa: BLE001
                out[url] = (False, None, f"fetch crash: {e!r}")
    return out


def validate_file(
    path: Path,
    issues_by_slug: dict[str, dict[str, Any]],
    candidates_by_slug: dict[str, dict[str, Any]],
) -> ValidationResult:
    payload = json.loads(path.read_text())
    slug = payload.get("candidate_slug") or path.stem
    res = ValidationResult(candidate_slug=slug)

    cand = candidates_by_slug.get(slug)
    if not cand:
        res.errors.append(f"unknown candidate slug: {slug}")
        return res
    res.candidate_id = cand["id"]

    positions = payload.get("positions") or []
    if not isinstance(positions, list):
        res.errors.append("positions must be a list")
        return res

    urls = collect_urls(payload)
    cache = fetch_all(urls)
    failed_urls = [u for u, (ok, _, msg) in cache.items() if not ok]
    if failed_urls:
        res.rejected_sources.extend(failed_urls)

    seen_issue_slugs: set[str] = set()
    for entry in positions:
        issue_slug = (entry or {}).get("issue_slug")
        if issue_slug not in issues_by_slug:
            res.errors.append(f"unknown issue slug: {issue_slug!r}")
            continue
        if issue_slug in seen_issue_slugs:
            res.errors.append(f"duplicate issue entry: {issue_slug}")
            continue
        seen_issue_slugs.add(issue_slug)

        cleaned, errs = validate_position_entry(entry, issue_slug, cache)
        res.errors.extend(errs)
        if cleaned is None:
            continue

        res.positions_to_apply.append(
            {
                "issue_id": issues_by_slug[issue_slug]["id"],
                "issue_slug": issue_slug,
                **cleaned,
            }
        )
        if cleaned["stance"] == "no_mention":
            res.no_mention_count += 1
        else:
            res.coded_count += 1

    missing = set(issues_by_slug.keys()) - seen_issue_slugs
    for m in missing:
        res.errors.append(f"missing position entry for issue: {m}")

    return res


# ----------------------------------------------------------------------
# Stage into trackerv3.xlsx (Positions v2 + Sources sheets)


def stage_into_xlsx(
    results: list[ValidationResult],
    candidates: list[dict[str, Any]],
) -> tuple[int, int, int]:
    """Returns (positions_updated, sources_removed, sources_added)."""
    from openpyxl import load_workbook  # type: ignore[import-not-found]
    from datetime import datetime as _datetime

    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"xlsx not found at {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH)
    pos_ws = wb["Positions v2"]
    src_ws = wb["Sources"]
    cand_ws = wb["Candidates"]

    # Map candidate (name, state) -> xlsx canonical id (e.g. "cotton-thomas").
    cand_headers = [c.value for c in cand_ws[1]]
    cand_col = {name: idx + 1 for idx, name in enumerate(cand_headers)}
    xlsx_id_by_name_state: dict[tuple[str, str], str] = {}
    for ridx in range(2, cand_ws.max_row + 1):
        xid = cand_ws.cell(row=ridx, column=cand_col["id"]).value
        nm = cand_ws.cell(row=ridx, column=cand_col["name"]).value
        st = cand_ws.cell(row=ridx, column=cand_col["state"]).value
        if not (xid and nm and st):
            continue
        xlsx_id_by_name_state[(str(nm).strip(), str(st).strip().upper())] = str(
            xid
        ).strip()

    # Make sure reviewStatus column exists.
    pos_headers = [c.value for c in pos_ws[1]]
    if "reviewStatus" not in pos_headers:
        new_col = len(pos_headers) + 1
        pos_ws.cell(row=1, column=new_col).value = "reviewStatus"
        pos_headers = [c.value for c in pos_ws[1]]
    pos_col = {name: idx + 1 for idx, name in enumerate(pos_headers)}
    required_pos = {"id", "candidateId", "topicId", "stance", "confidence",
                    "summary", "lastUpdated", "reviewStatus"}
    missing = required_pos - set(pos_col)
    if missing:
        raise ValueError(f"Positions v2 missing columns: {missing}")

    src_headers = [c.value for c in src_ws[1]]
    src_col = {name: idx + 1 for idx, name in enumerate(src_headers)}

    # Index Positions v2 by (xlsx_candidateId, topicId).
    pos_row_by_key: dict[tuple[str, str], tuple[int, str]] = {}
    for ridx in range(2, pos_ws.max_row + 1):
        cid = pos_ws.cell(row=ridx, column=pos_col["candidateId"]).value
        tid = pos_ws.cell(row=ridx, column=pos_col["topicId"]).value
        pid = pos_ws.cell(row=ridx, column=pos_col["id"]).value
        if not (cid and tid and pid):
            continue
        pos_row_by_key[(str(cid).strip(), str(tid).strip())] = (
            ridx,
            str(pid).strip(),
        )

    cand_by_id = {c["id"]: c for c in candidates}

    affected_position_ids: set[str] = set()
    pos_updates = 0

    for res in results:
        if not res.candidate_id:
            continue
        cand = cand_by_id.get(res.candidate_id)
        if not cand:
            print(
                f"  warn: no candidate record for {res.candidate_slug}",
                file=sys.stderr,
            )
            continue
        xlsx_cid = xlsx_id_by_name_state.get((cand["name"], cand["state"]))
        if not xlsx_cid:
            print(
                f"  warn: no xlsx Candidates row for {res.candidate_slug}",
                file=sys.stderr,
            )
            continue

        for p in res.positions_to_apply:
            match = pos_row_by_key.get((xlsx_cid, p["issue_slug"]))
            if not match:
                print(
                    f"  warn: no Positions v2 row for ({xlsx_cid}, {p['issue_slug']})",
                    file=sys.stderr,
                )
                continue
            ridx, pid = match
            affected_position_ids.add(pid)

            xlsx_stance = STANCE_TO_XLSX[p["stance"]]
            xlsx_conf = (
                CONFIDENCE_TO_XLSX.get(p["confidence"] or "low", "Low")
                if p["stance"] != "no_mention"
                else "N/A"
            )
            pos_ws.cell(row=ridx, column=pos_col["stance"]).value = xlsx_stance
            pos_ws.cell(row=ridx, column=pos_col["confidence"]).value = xlsx_conf
            pos_ws.cell(row=ridx, column=pos_col["summary"]).value = p["summary"]
            pos_ws.cell(row=ridx, column=pos_col["lastUpdated"]).value = (
                _datetime.fromisoformat(TODAY)
            )
            pos_ws.cell(row=ridx, column=pos_col["reviewStatus"]).value = "draft"
            pos_updates += 1

    # Sources: delete-then-append for affected positionIds.
    rows_to_delete: list[int] = []
    for ridx in range(2, src_ws.max_row + 1):
        pid = src_ws.cell(row=ridx, column=src_col["positionId"]).value
        if pid and str(pid).strip() in affected_position_ids:
            rows_to_delete.append(ridx)
    for ridx in sorted(rows_to_delete, reverse=True):
        src_ws.delete_rows(ridx, 1)
    sources_removed = len(rows_to_delete)

    sources_added = 0
    for res in results:
        if not res.candidate_id:
            continue
        cand = cand_by_id.get(res.candidate_id)
        if not cand:
            continue
        xlsx_cid = xlsx_id_by_name_state.get((cand["name"], cand["state"]))
        if not xlsx_cid:
            continue
        for p in res.positions_to_apply:
            match = pos_row_by_key.get((xlsx_cid, p["issue_slug"]))
            if not match:
                continue
            _, pid = match
            for s in p["sources"]:
                src_ws.append([None] * len(src_headers))
                ridx = src_ws.max_row
                src_ws.cell(row=ridx, column=src_col["positionId"]).value = pid
                src_ws.cell(
                    row=ridx, column=src_col["type"]
                ).value = SOURCE_TYPE_TO_XLSX.get(s["type"], s["type"])
                src_ws.cell(row=ridx, column=src_col["title"]).value = s.get("title")
                src_ws.cell(row=ridx, column=src_col["url"]).value = s["url"]
                d = s.get("date")
                if d:
                    try:
                        src_ws.cell(row=ridx, column=src_col["date"]).value = (
                            _datetime.fromisoformat(d[:10])
                        )
                    except ValueError:
                        pass
                src_ws.cell(row=ridx, column=src_col["excerpt"]).value = s.get("excerpt")
                sources_added += 1

    wb.save(XLSX_PATH)
    return pos_updates, sources_removed, sources_added


# ----------------------------------------------------------------------
# Reporting


def write_report(results: list[ValidationResult], applied_count: int, dry_run: bool) -> None:
    lines = [
        "# Research merge report",
        "",
        f"- Run date: {TODAY}",
        f"- Mode: {'dry-run (no writes)' if dry_run else 'applied'}",
        f"- Candidate files processed: {len(results)}",
        f"- Position rows {'would be' if dry_run else ''} updated: {applied_count}",
        "",
    ]
    for res in sorted(results, key=lambda r: r.candidate_slug):
        coded = res.coded_count
        nm = res.no_mention_count
        bad = len(res.rejected_sources) + len(res.errors)
        lines.append(f"## {res.candidate_slug}")
        lines.append(f"- coded: {coded}, no_mention: {nm}, issues w/ errors: {bad}")
        if res.rejected_sources:
            lines.append(f"- rejected URLs ({len(res.rejected_sources)}):")
            for u in res.rejected_sources[:10]:
                lines.append(f"  - {u}")
            if len(res.rejected_sources) > 10:
                lines.append(f"  - … +{len(res.rejected_sources) - 10} more")
        if res.errors:
            lines.append(f"- errors ({len(res.errors)}):")
            for e in res.errors[:20]:
                lines.append(f"  - {e}")
            if len(res.errors) > 20:
                lines.append(f"  - … +{len(res.errors) - 20} more")
        lines.append("")
    REPORT_PATH.write_text("\n".join(lines) + "\n")


# ----------------------------------------------------------------------
# Entrypoint


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "slug", nargs="?", help="Optional: limit to one candidate slug"
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not XLSX_PATH.exists():
        print(f"error: {XLSX_PATH} not found", file=sys.stderr)
        return 1

    issues = json.loads(ISSUES_PATH.read_text())
    issues_by_slug = {i["slug"]: i for i in issues}
    candidates = json.loads(CANDIDATES_PATH.read_text())
    candidates_by_slug = {c["slug"]: c for c in candidates}

    if not CACHE.exists():
        print(f"error: {CACHE} not found — run research first", file=sys.stderr)
        return 1

    cache_files = sorted(
        f for f in CACHE.glob("*.json") if not f.name.startswith("_")
    )
    if args.slug:
        cache_files = [f for f in cache_files if f.stem == args.slug]
        if not cache_files:
            print(f"error: no cached file for slug {args.slug}", file=sys.stderr)
            return 1

    print(f"Validating {len(cache_files)} candidate file(s)…")
    results: list[ValidationResult] = []
    for f in cache_files:
        print(f"  {f.name} …", end=" ", flush=True)
        res = validate_file(f, issues_by_slug, candidates_by_slug)
        results.append(res)
        print(
            f"coded {res.coded_count}, no_mention {res.no_mention_count}, "
            f"rejected URLs {len(res.rejected_sources)}, errors {len(res.errors)}"
        )

    applied = 0
    if not args.dry_run:
        pos_updates, src_removed, src_added = stage_into_xlsx(results, candidates)
        applied = pos_updates
        print(
            f"\nStaged into {XLSX_PATH.relative_to(ROOT.parent)}:"
            f" {pos_updates} positions (reviewStatus=draft),"
            f" {src_removed} sources removed,"
            f" {src_added} sources appended."
        )
        print(
            "Next step: open the xlsx, review the draft rows, set"
            " reviewStatus to 'approved' on rows you accept, then run"
            " python3 scripts/build_tracker_json.py to publish to the site."
        )
    else:
        applied = sum(len(r.positions_to_apply) for r in results)
        print(f"\n[dry-run] would stage {applied} position-row updates as draft")

    write_report(results, applied, args.dry_run)
    print(f"Report: {REPORT_PATH.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
